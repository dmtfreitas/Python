
# coding: UTF-8

# format_sheet_with_last_updated.py

from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook ('Fracht_Yload.xlsx')
ws = wb['Sheet1']

ws['AH1'] = 'LAST UPDATED'
ws['AH2'] = datetime.now ().strftime ('%Y-%m-%d %H:%M:%S')

wb.save ('Fracht_Yload_LAST_UPDATED.xlsx')